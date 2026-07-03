"""
Sincroniza la caché local de ocupación (para validar huéspedes).

Uso:
    python sincronizar_ocupacion.py                  → descarga desde Google Sheets
    python sincronizar_ocupacion.py archivo.xlsx     → importa desde un Excel local
"""
import csv
import io
import sys

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

import ocupacion


def importar_excel(ruta):
    import openpyxl
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb[wb.sheetnames[0]]

    salida = io.StringIO()
    writer = csv.writer(salida, lineterminator='\n')
    for fila in ws.iter_rows(values_only=True):
        writer.writerow(['' if c is None else str(c) for c in fila])

    ocupacion.guardar_cache(salida.getvalue(), 'excel')
    print(f'📄 Importado desde Excel: {ruta}')


def main():
    if len(sys.argv) > 1:
        importar_excel(sys.argv[1])
    else:
        try:
            texto = ocupacion._descargar_google()
            ocupacion.guardar_cache(texto, 'google')
            print('☁️  Descargado desde Google Sheets.')
        except Exception as e:
            print(f'❌ No se pudo descargar de Google Sheets: {e}')
            print('   Puedes importar un Excel local: python sincronizar_ocupacion.py archivo.xlsx')
            sys.exit(1)

    huespedes = ocupacion.obtener_huespedes(forzar=False) or []
    # Releer directo de la caché recién escrita para mostrar el resumen real
    texto, _ = ocupacion._leer_cache()
    huespedes = ocupacion._parsear_csv(texto)

    print(f'✅ {len(huespedes)} huéspedes con cédula registrados.')
    if huespedes:
        ej = huespedes[0]
        print(f'   Ejemplo: hab {ej["habitacion"]} (base {ej["hab_base"]}) — '
              f'cédula {ej["cedula"]} — {ej["nombre"] or "sin nombre"}')


if __name__ == '__main__':
    main()
